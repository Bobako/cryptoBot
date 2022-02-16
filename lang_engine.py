import socket

import openpyxl

from google.oauth2 import service_account
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from googleapiclient.discovery import build
import pprint
import io

FORMAT_DICT = {
    'Icon': "🇬🇧",
    'ChooseLang': "Please, select your language:",
    'LangUpdated': "Your language was set to english",
    'ChooseFiat': "Please select your fiat currency:",
    "FiatUpdated": "Your fiat currency was set to {}",
    "MainMenu":
        "Welcome back {}. \n\n"
        "How we can help you?",
    "Back": "Back to menu",
    "Operations": "Exchange",
    "Wallet": "Wallet",
    "Settings": "Settings",
    "Support": "Info & Support",
    "OperationsMenu":
        "What would you like to do?\n"
        "• Fast Exchange - Just  choose what you want to give and get & make your trade\n"
        "• Escrow Exchange - Exchange crypto with a person chosen by you\n"
        "• P2P - Make and take orders to trade crypto\n",
    "FastExchange": "Fast Exchange",
    "FastExchangeRate": "Exchange rate for these currencies:\n{}",
    "ExchangeResult": "You will get {} {}.\nConfirm?",
    "EnterSellSum": "Please enter sum to sell:",
    "EscrowExchange": "Escrow Exchange",
    "EnterUsername": "Enter username:",
    "NoSuchUser": "There is no such user in system, retry",
    "Counterparty": "Counterparty",
    "CPNotified": "We notified counterparty and wait for response",
    "CPNotificationChat": "You got new escrow exchange offer from @{}:",
    "CPNotificationNoChat": "You got new escrow exchange offer from @{}:",
    "CPNotificationOrder": "You got new escrow exchange offer from @{}:",
    "EscrowDescription": "@{} wants to {}",
    'Accept': "Accept",
    'Deny': "Deny",
    "Accepted": "Your escrow exchange offer was accepted:\n{}",
    "Invite": "\nNow join this chat with your counterparty and moderator:\n{}",
    "EscrowInstructions": "Now @{1} can transfer funds to @{0} account. After transfer @{0} may confirm receipt by /confirm.\n"
                          "/cancel to cancel escrow",
    "EscrowFinished": "Escrow finished! Now rate @{}",
    "FirstExchange": "This user did not perform the exchange",
    "UserRate": "{} has {} finished exchanges, average rate: {}",
    "Thanks": "Thanks!",
    "P2P": "P2P Exchange",
    "P2PMenu": "What would you like to do?",
    "SeeOrders": "View active orders",
    "OrdersView": "Orders page {} out of {}",
    "CreateOrder": "Create order",
    "EnterMinBuy": "Enter min price for your offer",
    "OrderCreated": "Order was created",
    "SeeOwnOrders": "View your orders",
    "DeleteOrder": "{}\nDo you want to delete order?",
    "OrderDeleted": "Order deleted",
    "OfferForOrder": "Enter your offer sum (at least {}):",
    "OrderAccepted": "Your {} order was accepted. @{} offers {} {}",
    "OrderFormat": "@{} sell {} {} for at least {} {}",
    "NoOrders": "There is not orders now",
    "WalletMenu":
        "What would you like to do?\n"
        "• Deposit money on your account in the bot\n"
        "• Withdraw money from the bot to your personal account\n"
        "• Balance - see your balance",
    "Deposit": "Deposit",
    "Withdraw": "Withdraw",
    "Balance": "Balance",
    "NoBalance": "You hadnt deposit any currencies",
    "CurrentBalance": "Now you have {} {}",
    "OperationAmount": "Please enter {} sum",
    "MinDeposit": "Minimal deposit amount is {} {}",
    "ToSmallAmount": "To small amount. Enter at least {} {}",
    "FormatError": "Message you sent isnt a correct amount, retry",
    "OperationCurrency": "Please choose a currency to {}",
    "AccountAddress": "Make a transaction to this address:\n",
    "CryptoDeposit": "Due to anonymous crypto transactions, transfer exactly {},"
                     " so that we can identify your transfer ",
    "TransferConfirmation": "After transfer press the confirm button",
    "OperationCancelled": "Operation was cancelled",
    "Confirm": "Confirm",
    "OperationProceed": "Operation is being proceed. Operation got id #{}",
    "OperationConfirmed": "Your {} confirmed",
    "OperationRejected": "Your {} rejected",
    "CurrencyChosen": "{} chosen.",
    "AddressToWithdraw": "Please enter a wallet address to withdraw",
    "NotEnough": "You dont have enough on your balance",
    "SettingsMenu": "Settings:",
    'ChangeLanguage': "Change language",
    'ChangeFiat': "Change your fiat currency",

    'SupportMenu': "What would you like to see?",
    'Ask': "Ask",
    'AskLink': "www.example.com",
    'JoinCommunity': "Join community",
    'CommunityLink': "www.example.com",
    'Fees': "Fees",
    'FeesText': "Sample Fees text",
    'Rates': "Exchange rates",
    "Buy": "Buy",
    "Sell": "Sell",
    "CryptoDepositAddress": "",
    "EscrowExplanation": "",
    "ChatEscrow": "",
    "NoChatEscrow": "",
    "SelectOrderSellCur": "",
    "SelectOrderBuyCur": "",
    "YourOffer": "YourOffer",
    "DeleteEscrow": "",
    "NoOffers": "",
    "MinTradeAmount": "",
    "YourRate": "",
    "ToBigAmount": ""
}


def load_lang_table(filename="langs.xlsx"):
    headers = list(FORMAT_DICT.keys())
    wb = openpyxl.load_workbook(filename)
    ws = wb["langs"]
    column = 3
    langs = {}
    while ws.cell(row=1, column=column).value:
        lang_name = ws.cell(row=1, column=column).value
        lang = {}
        for i, header in enumerate(headers):
            row = i + 2
            lang[header] = ws.cell(row, column).value
        langs[lang_name] = lang
        column += 1
    return langs


def create_lang_table(filename="langs.xlsx"):
    headers = list(FORMAT_DICT.keys())
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, header in enumerate(headers):
        row = i + 2
        ws.cell(row=row, column=2).value = header
        ws.cell(row=row, column=3).value = FORMAT_DICT[header]
    wb.save(filename)


def load_langs_table_from_google(file_name="langs.xlsx"):
    print("Загрузка языковых таблиц...", end="")
    scopes = ['https://www.googleapis.com/auth/drive']
    service_account_file = './pythonsheets-331614-0ebcc874bffa.json'
    credentials = service_account.Credentials.from_service_account_file(
        service_account_file, scopes=scopes)
    service = build('drive', 'v3', credentials=credentials)
    try:
        results = service.files().list(pageSize=10,
                                   fields="nextPageToken, files(id, name, mimeType)").execute()
    except socket.timeout:
        print("Ошибка (таймаут от серверов гугла). Будет использоваться старая версия.")
    else:
        file_id = ''
        for file in results["files"]:
            if file["name"] == file_name:
                file_id = file["id"]
                break

        request = service.files().get_media(fileId=file_id)
        filename = file_name
        fh = io.FileIO(filename, 'wb')
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
            print(".", end="")
        print("готово\n")
    return load_lang_table(file_name)


def load_reqs(filename="langs.xlsx"):
    wb = openpyxl.load_workbook(filename)
    ws = wb["reqs"]
    row = 1
    reqs = {}
    while ws.cell(row=row, column=1).value:
        key = ws.cell(row=row, column=1).value
        value = ws.cell(row=row, column=2).value
        reqs[key] = value
        row += 1
    wb.close()
    return reqs


if __name__ == '__main__':
    pass
    # print(load_langs_table_from_google())
